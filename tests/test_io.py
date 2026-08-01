import os
import sys
import numpy as np
import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dsm_optimizer.io.reader import read_dsm
from dsm_optimizer.io.writer import write_excel

SAMPLE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "server/sample_dsm.xlsx")


def test_read_sample_xlsx():
    assert os.path.exists(SAMPLE), "sample_dsm.xlsx missing from repo root"
    matrix, labels = read_dsm(SAMPLE)
    assert len(labels) > 0
    assert matrix.shape == (len(labels), len(labels))
    assert np.all(np.diag(matrix) == 0)


def test_read_csv_roundtrip(tmp_path):
    csv_path = tmp_path / "mini.csv"
    csv_path.write_text("label,A,B,C\nA,,1,0\nB,0,,1\nC,1,0,\n")
    matrix, labels = read_dsm(str(csv_path))
    assert labels == ["A", "B", "C"]
    assert matrix[0, 1] == 1
    assert matrix[1, 2] == 1
    assert matrix[2, 0] == 1


def test_write_excel_produces_readable_file(tmp_path):
    out_path = tmp_path / "out.xlsx"
    matrix = np.array([[0, 1, 0], [1, 0, 1], [0, 1, 0]], dtype=float)
    labels = ["X", "Y", "Z"]
    clusters = [0, 0, 1]
    write_excel(str(out_path), matrix, labels, clusters, title="Test DSM")
    assert out_path.exists()

    wb = openpyxl.load_workbook(str(out_path))
    ws = wb.active
    assert ws.cell(1, 1).value == "Test DSM"
    # Column headers start at col 2 (B) and should match labels in order
    written_labels = [ws.cell(1, j + 2).value for j in range(len(labels))]
    assert written_labels == labels
