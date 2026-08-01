import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_CLUSTER_COLORS = [
    "AED6F1", "A9DFBF", "F9E79F", "F5CBA7", "D7BDE2",
    "FAD7A0", "A3E4D7", "85C1E9", "D5DBDB", "FADBD8",
]
_DIAG = "BFC9CA"
_EXTERNAL = "F1948A"   # red for inter-cluster marks
_THIN = Side(style='thin', color='888888')
_THICK = Side(style='medium', color='000000')


def write_excel(filepath, matrix, labels, clusters, title="Optimized DSM"):
    n = len(labels)
    unique = sorted(set(clusters))
    color_map = {c: _CLUSTER_COLORS[i % len(_CLUSTER_COLORS)] for i, c in enumerate(unique)}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Optimized DSM"

    # ── Title cell ────────────────────────────────────────────────────────────
    ws.cell(1, 1, title).font = Font(bold=True, size=10)

    # ── Column headers ────────────────────────────────────────────────────────
    for j, lbl in enumerate(labels):
        c = ws.cell(1, j + 2, lbl)
        c.font = Font(bold=True, size=7)
        c.alignment = Alignment(horizontal='center', textRotation=90)
        c.fill = PatternFill("solid", fgColor=color_map[clusters[j]])

    # ── Rows ──────────────────────────────────────────────────────────────────
    for i in range(n):
        # Row label
        rl = ws.cell(i + 2, 1, labels[i])
        rl.font = Font(bold=True, size=8)
        rl.alignment = Alignment(vertical='center')
        rl.fill = PatternFill("solid", fgColor=color_map[clusters[i]])

        for j in range(n):
            cell = ws.cell(i + 2, j + 2)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
            val = matrix[i][j]

            if i == j:
                cell.fill = PatternFill("solid", fgColor=_DIAG)
            elif val > 0:
                cell.value = int(val) if val == int(val) else round(val, 2)
                if clusters[i] == clusters[j]:
                    cell.fill = PatternFill("solid", fgColor=color_map[clusters[i]])
                else:
                    cell.fill = PatternFill("solid", fgColor=_EXTERNAL)

    # ── Thick cluster borders ─────────────────────────────────────────────────
    for c in unique:
        pos = [i for i, cl in enumerate(clusters) if cl == c]
        if not pos:
            continue
        r0, r1 = min(pos) + 2, max(pos) + 2
        c0, c1 = min(pos) + 2, max(pos) + 2
        for row in range(r0, r1 + 1):
            for col in range(c0, c1 + 1):
                cell = ws.cell(row, col)
                l = _THICK if col == c0 else cell.border.left
                r = _THICK if col == c1 else cell.border.right
                t = _THICK if row == r0 else cell.border.top
                b = _THICK if row == r1 else cell.border.bottom
                cell.border = Border(left=l, right=r, top=t, bottom=b)

    # ── Column/row sizes ──────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 22
    for j in range(n):
        ws.column_dimensions[get_column_letter(j + 2)].width = 3.5
    for i in range(n + 1):
        ws.row_dimensions[i + 1].height = 13

    wb.save(filepath)
    return filepath
